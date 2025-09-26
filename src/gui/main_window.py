import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import json
from ..core import get_localization, t

class TimetableApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # Initialize localization
        self.localization = get_localization()
        
        self.title(t("app_title"))
        self.geometry("1200x800")
        
        # Database path
        self.db_path = "data/database/school_timetable.db"

        # --- Style ---
        self.style = ttk.Style(self)
        self.style.theme_use('clam') # 'clam', 'alt', 'default', 'classic'
        
        # Get localized font
        font_family = self.localization.get_font_family()
        
        self.style.configure("TButton", padding=6, relief="flat", background="#cceeff", borderwidth=0)
        self.style.map("TButton",
            foreground=[('pressed', 'red'), ('active', 'blue')],
            background=[('pressed', '!disabled', 'black'), ('active', 'white')]
        )
        self.style.configure("TFrame", background="#f0f0f0")
        
        # Additional styles for dialog elements with localized font
        self.style.configure("Large.TRadiobutton", font=(font_family, 10, 'bold'))
        self.style.configure("Custom.TRadiobutton", font=(font_family, 10))
        self.style.configure("Bold.TRadiobutton", font=(font_family, 10, 'bold'))
        self.style.configure("Header.TLabel", background="#007acc", foreground="white", padding=10, font=(font_family, 12, 'bold'))
        self.style.configure("Slot.TButton", font=(font_family, 9), width=15)
        self.style.configure("Time.TLabel", background="#f0f0f0", font=(font_family, 10, 'bold'))
        
        # Configure for RTL support if needed
        if self.localization.is_rtl():
            self.style.configure("RTL.TLabel", anchor="e")
            self.style.configure("RTL.TButton", anchor="e")

        # --- Menu Bar ---
        self.create_menu()

        # --- Main Layout ---
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Controls Frame ---
        controls_frame = ttk.Frame(main_frame, padding="10")
        controls_frame.pack(fill=tk.X)
        
        # Adjust layout based on RTL
        view_side = tk.RIGHT if self.localization.is_rtl() else tk.LEFT
        button_side = tk.LEFT if self.localization.is_rtl() else tk.RIGHT
        
        font_family = self.localization.get_font_family()
        ttk.Label(controls_frame, text=t("view_label"), font=(font_family, 11, 'bold')).pack(side=view_side, padx=5)
        
        self.view_var = tk.StringVar(value="Classes")
        class_view_btn = ttk.Radiobutton(controls_frame, text=t("classes"), variable=self.view_var, value="Classes", command=self.on_view_change)
        teacher_view_btn = ttk.Radiobutton(controls_frame, text=t("teachers"), variable=self.view_var, value="Teachers", command=self.on_view_change)
        class_view_btn.pack(side=view_side, padx=5)
        teacher_view_btn.pack(side=view_side, padx=5)
        
        self.item_selector = ttk.Combobox(controls_frame, state="readonly", width=20)
        self.item_selector.pack(side=view_side, padx=20)
        self.item_selector.bind("<<ComboboxSelected>>", self.draw_timetable)

        generate_btn = ttk.Button(controls_frame, text=t("generate_schedule"), command=self.generate_schedule)
        generate_btn.pack(side=button_side, padx=5)

        export_excel_btn = ttk.Button(controls_frame, text=t("export") + " Excel", command=self.export_excel)
        export_excel_btn.pack(side=button_side, padx=5)
        
        export_pdf_btn = ttk.Button(controls_frame, text=t("export") + " PDF", command=self.export_pdf)
        export_pdf_btn.pack(side=button_side, padx=5)


        # --- Timetable Frame ---
        self.timetable_frame = ttk.Frame(main_frame)
        self.timetable_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.ensure_database_exists()
        self.load_initial_data()
        self.draw_timetable()

    def ensure_database_exists(self):
        """Ensure the database exists with sample data"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            # Check if tables exist
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='teachers'")
            if not cursor.fetchone():
                # Database doesn't exist, create it
                conn.close()
                from ..database.database_setup import create_connection, create_tables, add_sample_data
                conn = create_connection()
                if conn:
                    create_tables(conn)
                    add_sample_data(conn)
                    conn.close()
                    print("Database created with sample data")
            else:
                conn.close()
        except Exception as e:
            print(f"Error ensuring database exists: {e}")

    def load_initial_data(self):
        """Loads data into the combobox based on the view selected"""
        view = self.view_var.get()
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        if view == "Classes":
            cursor.execute("SELECT name FROM classes ORDER BY name")
        else:
            cursor.execute("SELECT name FROM teachers ORDER BY name")
        
        items = [row[0] for row in cursor.fetchall()]
        self.item_selector['values'] = items
        if items:
            self.item_selector.set(items[0])
        conn.close()

    def on_view_change(self):
        """Handle view change between Classes and Teachers"""
        self.load_initial_data()  # Update dropdown contents
        self.draw_timetable()     # Redraw timetable with new selection


    def draw_timetable(self, event=None):
        """Clears and redraws the timetable grid."""
        for widget in self.timetable_frame.winfo_children():
            widget.destroy()

        # Define headers with localization
        days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]

        # Create Day Headers
        for i, day in enumerate(days):
            ttk.Label(self.timetable_frame, text=day, style="Header.TLabel", anchor="center").grid(row=0, column=i + 1, sticky="nsew", padx=1, pady=1)

        # Create Time Headers
        for i, time in enumerate(times):
            ttk.Label(self.timetable_frame, text=time, style="Time.TLabel", anchor="center").grid(row=i + 1, column=0, sticky="nsew", padx=5)

        # Get current selection
        selected_item = self.item_selector.get()
        view_type = self.view_var.get()
        
        # Load schedule data from database
        schedule_data = self.load_schedule_data(selected_item, view_type)

        # Create Timetable Slots (Buttons)
        for r, time in enumerate(times):
            for c, day in enumerate(days):
                # Look for lesson at this time slot
                lesson_info = ""
                slot_data = schedule_data.get((c, r), None)  # (day, period)
                
                if slot_data:
                    if view_type == "Classes":
                        lesson_info = f"{slot_data['subject']}\n{slot_data['teacher']}\n{slot_data['room']}"
                    else:  # Teachers view
                        lesson_info = f"{slot_data['subject']}\n{slot_data['class']}\n{slot_data['room']}"
                else:
                    lesson_info = t("free")
                
                slot_btn = ttk.Button(self.timetable_frame, text=lesson_info, style="Slot.TButton")
                slot_btn.grid(row=r + 1, column=c + 1, sticky="nsew", padx=1, pady=1)
                
                # Add click binding for editing (will implement later)
                slot_btn.bind("<Button-1>", lambda e, d=c, p=r: self.on_slot_click(d, p))

        # Configure grid weights to make it expand
        for i in range(len(days) + 1):
            self.timetable_frame.grid_columnconfigure(i, weight=1)
        for i in range(len(times) + 1):
            self.timetable_frame.grid_rowconfigure(i, weight=1)

    def load_schedule_data(self, selected_item, view_type):
        """Load schedule data for the selected item from database"""
        if not selected_item:
            return {}
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        schedule_data = {}
        
        if view_type == "Classes":
            # Load schedule for specific class
            cursor.execute("""
                SELECT s.day_of_week, s.timeslot, sub.name as subject, t.name as teacher, r.name as room
                FROM schedules s
                JOIN classes c ON s.class_id = c.id
                JOIN subjects sub ON s.subject_id = sub.id
                JOIN teachers t ON s.teacher_id = t.id
                JOIN rooms r ON s.room_id = r.id
                WHERE c.name = ?
                ORDER BY s.day_of_week, s.timeslot
            """, (selected_item,))
        else:  # Teachers view
            # Load schedule for specific teacher
            cursor.execute("""
                SELECT s.day_of_week, s.timeslot, sub.name as subject, c.name as class, r.name as room
                FROM schedules s
                JOIN classes c ON s.class_id = c.id
                JOIN subjects sub ON s.subject_id = sub.id
                JOIN teachers t ON s.teacher_id = t.id
                JOIN rooms r ON s.room_id = r.id
                WHERE t.name = ?
                ORDER BY s.day_of_week, s.timeslot
            """, (selected_item,))
        
        results = cursor.fetchall()
        conn.close()
        
        for row in results:
            day, period = row[0], row[1]
            if view_type == "Classes":
                schedule_data[(day, period)] = {
                    'subject': row[2],
                    'teacher': row[3],
                    'room': row[4]
                }
            else:
                schedule_data[(day, period)] = {
                    'subject': row[2],
                    'class': row[3],
                    'room': row[4]
                }
        
        return schedule_data

    def on_slot_click(self, day, period):
        """Handle click on a timetable slot for editing"""
        selected_item = self.item_selector.get()
        view_type = self.view_var.get()
        
        if not selected_item:
            return
        
        schedule_data = self.load_schedule_data(selected_item, view_type)
        slot_data = schedule_data.get((day, period), None)
        
        # Open edit dialog
        self.open_slot_edit_dialog(day, period, selected_item, view_type, slot_data)

    def open_slot_edit_dialog(self, day, period, selected_item, view_type, current_data):
        """Open dialog to edit a time slot"""
        days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
        
        window = tk.Toplevel(self)
        window.title(f"{t('edit')} {days[day]} {times[period]}")
        window.geometry("400x300")
        
        # Apply RTL layout if Arabic
        font_family = self.localization.get_font_family()
        if self.localization.is_rtl():
            window.option_add('*TLabel*font', font_family)
        
        ttk.Label(window, text=f"{t('editing')}: {selected_item}", 
                 font=(font_family, 10)).pack(pady=10)
        ttk.Label(window, text=f"{t('time')}: {days[day]} {times[period]}",
                 font=(font_family, 10)).pack(pady=5)
        
        if current_data:
            ttk.Label(window, text=t("current_assignment"), 
                     font=(font_family, 10, 'bold')).pack(pady=10)
            if view_type == "Classes":
                ttk.Label(window, text=f"{t('subject')}: {current_data['subject']}",
                         font=(font_family, 9)).pack()
                ttk.Label(window, text=f"{t('teacher')}: {current_data['teacher']}",
                         font=(font_family, 9)).pack()
                ttk.Label(window, text=f"{t('room')}: {current_data['room']}",
                         font=(font_family, 9)).pack()
            else:
                ttk.Label(window, text=f"{t('subject')}: {current_data['subject']}",
                         font=(font_family, 9)).pack()
                ttk.Label(window, text=f"{t('class')}: {current_data['class']}",
                         font=(font_family, 9)).pack()
                ttk.Label(window, text=f"{t('room')}: {current_data['room']}",
                         font=(font_family, 9)).pack()
        
        # Action buttons
        button_frame = ttk.Frame(window)
        button_frame.pack(pady=20)
        
        if current_data:
            ttk.Button(button_frame, text=t("remove_lesson"), 
                      command=lambda: self.remove_lesson(day, period, selected_item, view_type, window)).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text=t("move_lesson"), 
                  command=lambda: self.move_lesson_dialog(day, period, selected_item, view_type, current_data, window)).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text=t("close"), 
                  command=window.destroy).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Cancel", command=window.destroy).pack(side=tk.LEFT, padx=5)

    def remove_lesson(self, day, period, selected_item, view_type, parent_window):
        """Remove a lesson from the schedule"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            if view_type == "Classes":
                cursor.execute("""
                    DELETE FROM schedules 
                    WHERE class_id = (SELECT id FROM classes WHERE name = ?) 
                    AND day_of_week = ? AND timeslot = ?
                """, (selected_item, day, period))
            else:
                cursor.execute("""
                    DELETE FROM schedules 
                    WHERE teacher_id = (SELECT id FROM teachers WHERE name = ?) 
                    AND day_of_week = ? AND timeslot = ?
                """, (selected_item, day, period))
            
            conn.commit()
            conn.close()
            
            parent_window.destroy()
            self.draw_timetable()
            print("Lesson removed successfully")
            
        except Exception as e:
            print(f"Error removing lesson: {e}")

    def move_lesson_dialog(self, old_day, old_period, selected_item, view_type, lesson_data, parent_window):
        """Open dialog to move a lesson to a new time slot"""
        if not lesson_data:
            print("No lesson to move")
            return
        
        window = tk.Toplevel(self)
        window.title(t("move_lesson_title"))
        window.geometry("500x400")
        
        # Apply RTL layout if Arabic
        font_family = self.localization.get_font_family()
        if self.localization.is_rtl():
            window.option_add('*TLabel*font', font_family)
        
        ttk.Label(window, text=t("select_new_time_slot"), 
                 font=(font_family, 12, 'bold')).pack(pady=10)
        
        # Create a mini timetable for selection
        days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
        
        frame = ttk.Frame(window)
        frame.pack(padx=20, pady=20)
        
        # Headers
        for i, day in enumerate(days):
            ttk.Label(frame, text=day, font=(font_family, 10, 'bold')).grid(row=0, column=i + 1, padx=2, pady=2)
        
        for i, time in enumerate(times):
            ttk.Label(frame, text=time, font=(font_family, 9)).grid(row=i + 1, column=0, padx=2, pady=2, sticky="e")
        
        # Create buttons for each slot
        for r, time in enumerate(times):
            for c, day in enumerate(days):
                btn = ttk.Button(frame, text=t("select"), width=8,
                               command=lambda d=c, p=r: self.move_lesson(old_day, old_period, d, p, selected_item, view_type, lesson_data, window, parent_window))
                btn.grid(row=r + 1, column=c + 1, padx=1, pady=1)
                
                # Disable current slot
                if c == old_day and r == old_period:
                    btn.config(state="disabled", text=t("current"))
        
        ttk.Button(window, text=t("cancel"), command=window.destroy).pack(pady=10)

    def move_lesson(self, old_day, old_period, new_day, new_period, selected_item, view_type, lesson_data, move_window, edit_window):
        """Move a lesson to a new time slot"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Check for conflicts at the new time slot
            conflict_check = self.check_move_conflicts(new_day, new_period, lesson_data, cursor)
            
            if conflict_check:
                print(f"Cannot move lesson: {conflict_check}")
                return
            
            # Update the lesson time
            if view_type == "Classes":
                cursor.execute("""
                    UPDATE schedules 
                    SET day_of_week = ?, timeslot = ?
                    WHERE class_id = (SELECT id FROM classes WHERE name = ?) 
                    AND day_of_week = ? AND timeslot = ?
                """, (new_day, new_period, selected_item, old_day, old_period))
            else:
                cursor.execute("""
                    UPDATE schedules 
                    SET day_of_week = ?, timeslot = ?
                    WHERE teacher_id = (SELECT id FROM teachers WHERE name = ?) 
                    AND day_of_week = ? AND timeslot = ?
                """, (new_day, new_period, selected_item, old_day, old_period))
            
            conn.commit()
            conn.close()
            
            move_window.destroy()
            edit_window.destroy()
            self.draw_timetable()
            print("Lesson moved successfully")
            
        except Exception as e:
            print(f"Error moving lesson: {e}")

    def check_move_conflicts(self, day, period, lesson_data, cursor):
        """Check for conflicts when moving a lesson"""
        # Check if teacher is already busy
        if 'teacher' in lesson_data:
            cursor.execute("""
                SELECT COUNT(*) FROM schedules s
                JOIN teachers t ON s.teacher_id = t.id
                WHERE t.name = ? AND s.day_of_week = ? AND s.timeslot = ?
            """, (lesson_data['teacher'], day, period))
            
            if cursor.fetchone()[0] > 0:
                return f"Teacher {lesson_data['teacher']} is already busy at this time"
        
        # Check if class is already busy (if we know the class)
        if 'class' in lesson_data:
            cursor.execute("""
                SELECT COUNT(*) FROM schedules s
                JOIN classes c ON s.class_id = c.id
                WHERE c.name = ? AND s.day_of_week = ? AND s.timeslot = ?
            """, (lesson_data['class'], day, period))
            
            if cursor.fetchone()[0] > 0:
                return f"Class {lesson_data['class']} is already busy at this time"
        
        # Check if room is already busy
        if 'room' in lesson_data:
            cursor.execute("""
                SELECT COUNT(*) FROM schedules s
                JOIN rooms r ON s.room_id = r.id
                WHERE r.name = ? AND s.day_of_week = ? AND s.timeslot = ?
            """, (lesson_data['room'], day, period))
            
            if cursor.fetchone()[0] > 0:
                return f"Room {lesson_data['room']} is already busy at this time"
        
        return None  # No conflicts
            
    def generate_schedule(self):
        """Generate a new schedule with algorithm selection"""
        # Create algorithm selection dialog
        dialog = tk.Toplevel(self)
        dialog.title(t("select_algorithm"))
        dialog.geometry("600x650")  # Increased size
        dialog.resizable(False, False)
        dialog.grab_set()  # Make dialog modal
        
        # Center dialog on parent
        dialog.transient(self)
        self.update_idletasks()
        x = (self.winfo_x() + (self.winfo_width() // 2)) - 300
        y = (self.winfo_y() + (self.winfo_height() // 2)) - 325
        dialog.geometry(f"600x650+{x}+{y}")
        
        font_family = self.localization.get_font_family()
        ttk.Label(dialog, text=t("choose_algorithm"), 
                 font=(font_family, 14, 'bold')).pack(pady=20)
        
        # Algorithm selection
        algorithm_var = tk.StringVar(value="fast_greedy")
        
        # Get localized algorithm information
        algorithms = [
            ("ultra_fast", t("ultra_fast"), t("ultra_fast_desc")),
            ("smart_greedy", t("smart_greedy"), t("smart_greedy_desc")), 
            ("ml_inspired", t("ml_inspired"), t("ml_inspired_desc")),
            ("fast_greedy", t("fast_greedy"), t("fast_greedy_desc")),
            ("ortools", t("ortools"), t("ortools_desc")),
            ("simple", t("simple"), t("simple_desc"))
        ]
        
        # Create scrollable frame for algorithms
        canvas_frame = ttk.Frame(dialog)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        canvas = tk.Canvas(canvas_frame, height=350, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create radio buttons with descriptions
        for i, (value, title, description) in enumerate(algorithms):
            algo_frame = ttk.LabelFrame(scrollable_frame, text="", padding=10)
            algo_frame.pack(fill=tk.X, padx=5, pady=8)
            
            # Radio button with larger font and RTL support
            anchor_side = tk.E if self.localization.is_rtl() else tk.W
            radio_btn = ttk.Radiobutton(algo_frame, text=title, variable=algorithm_var, 
                                       value=value)
            radio_btn.pack(anchor=anchor_side)
            radio_btn.configure(style="Large.TRadiobutton")
            
            # Description with better formatting and localized font
            desc_lines = description.split('\n')
            for line in desc_lines:
                if line.strip():
                    desc_label = ttk.Label(algo_frame, text=line, font=(font_family, 9), 
                                         foreground='#666666')
                    desc_label.pack(anchor=anchor_side, padx=20, pady=1)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Separator line
        separator = ttk.Separator(dialog, orient='horizontal')
        separator.pack(fill=tk.X, padx=20, pady=15)
        
        # Buttons frame - make it more prominent
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        def run_selected_algorithm():
            algorithm = algorithm_var.get()
            dialog.destroy()
            self.run_scheduling_algorithm(algorithm)
        
        # Larger, More prominent buttons with localized text
        generate_btn = ttk.Button(btn_frame, text=t("generate"), 
                                 command=run_selected_algorithm)
        generate_btn.pack(side=tk.LEFT, padx=15, pady=10)
        generate_btn.configure(width=20)  # Make button wider
        
        cancel_btn = ttk.Button(btn_frame, text=t("cancel"), 
                               command=dialog.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=15, pady=10)
        cancel_btn.configure(width=15)  # Make button wider
        
        # Add keyboard shortcuts
        dialog.bind('<Return>', lambda e: run_selected_algorithm())
        dialog.bind('<Escape>', lambda e: dialog.destroy())
        
        # Focus on the dialog
        dialog.focus_set()
    
    def run_scheduling_algorithm(self, algorithm: str):
        """Run the selected scheduling algorithm using the SolverFactory"""
        from ..solvers import SolverFactory, SolverType
        
        # Show progress
        progress_window = tk.Toplevel(self)
        progress_window.title(t("generating_schedule"))
        progress_window.geometry("400x150")
        progress_window.grab_set()
        
        # Apply RTL layout if Arabic
        if self.localization.is_rtl():
            progress_window.option_add('*TLabel*font', self.localization.get_font_family())
        
        # Center the progress window
        progress_window.transient(self)
        self.update_idletasks()
        x = (self.winfo_x() + (self.winfo_width() // 2)) - 200
        y = (self.winfo_y() + (self.winfo_height() // 2)) - 75
        progress_window.geometry(f"400x150+{x}+{y}")
        
        ttk.Label(progress_window, text=f"{t('running_algorithm')}: {algorithm.replace('_', ' ').title()}...", 
                 font=(self.localization.get_font_family(), 12)).pack(pady=20)
        
        progress_bar = ttk.Progressbar(progress_window, mode='indeterminate')
        progress_bar.pack(pady=10, padx=20, fill=tk.X)
        progress_bar.start()
        
        status_label = ttk.Label(progress_window, text=t("initializing"), 
                                font=(self.localization.get_font_family(), 10))
        status_label.pack(pady=10)
        
        # Update display
        progress_window.update()
        
        try:
            # Map algorithm names to SolverType enum
            solver_map = {
                "ultra_fast": SolverType.ULTRA_FAST,
                "smart_greedy": SolverType.SMART_GREEDY,
                "ml_inspired": SolverType.ML_INSPIRED,
                "fast_greedy": SolverType.FAST_GREEDY,
                "ortools": SolverType.ORTOOLS,
                "simple": SolverType.SIMPLE
            }
            
            solver_type = solver_map.get(algorithm)
            if not solver_type:
                raise ValueError(f"Unknown algorithm: {algorithm}")
            
            status_label.config(text=t("running_algorithm_status"))
            progress_window.update()
            
            # Use the solver factory
            db_path = "data/database/school_timetable.db"
            result = SolverFactory.solve(solver_type, db_path)
            
            progress_bar.stop()
            progress_window.destroy()
            
            if result.success:
                success_msg = f"✅ {t('schedule_generated_successfully')}\n\n"
                success_msg += f"{t('algorithm')}: {result.algorithm.replace('_', ' ').title()}\n"
                success_msg += f"{t('time_taken')}: {result.time_taken:.2f} {t('seconds')}\n"
                success_msg += f"{t('lessons_scheduled')}: {result.lessons_count}\n\n"
                success_msg += t("schedule_displayed_main_window")
                
                tk.messagebox.showinfo(t("schedule_generated"), success_msg)
                
                # Refresh the timetable display
                self.draw_timetable()
            else:
                error_msg = f"❌ {t('failed_to_generate_schedule')}\n\n"
                error_msg += f"{t('algorithm')}: {result.algorithm.replace('_', ' ').title()}\n"
                error_msg += f"{t('time_taken')}: {result.time_taken:.2f} {t('seconds')}\n\n"
                if result.error:
                    error_msg += f"{t('error')}: {result.error}\n\n"
                error_msg += t("check_lesson_requirements_constraints")
                
                tk.messagebox.showerror(t("scheduling_failed"), error_msg)
        
        except Exception as e:
            progress_bar.stop()
            progress_window.destroy()
            tk.messagebox.showerror(t("error"), f"{t('an_error_occurred')}:\n{str(e)}")
            print(f"Error in scheduling: {e}")  # Debug info
        
    def create_menu(self):
        """Create the application menu bar"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        
        # Data menu
        data_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=t("menu_data"), menu=data_menu)
        data_menu.add_command(label=t("manage_teachers"), command=self.manage_teachers)
        data_menu.add_command(label=t("manage_classes"), command=self.manage_classes)
        data_menu.add_command(label=t("manage_subjects"), command=self.manage_subjects)
        data_menu.add_command(label=t("manage_rooms"), command=self.manage_rooms)
        data_menu.add_separator()
        data_menu.add_command(label=t("set_lesson_requirements"), command=self.manage_lessons)
        data_menu.add_command(label=t("teacher_preferences"), command=self.manage_teacher_preferences)
        data_menu.add_command(label=t("teacher_availability"), command=self.manage_teacher_availability)
        
        # Rules menu
        rules_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=t("menu_rules"), menu=rules_menu)
        rules_menu.add_command(label=t("scheduling_rules"), command=self.manage_scheduling_rules)
        rules_menu.add_command(label=t("constraint_settings"), command=self.manage_constraints)
        rules_menu.add_command(label=t("time_settings"), command=self.manage_time_settings)
        
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=t("menu_tools"), menu=tools_menu)
        tools_menu.add_command(label=t("database_statistics"), command=self.show_database_stats)
        tools_menu.add_command(label=t("clear_all_schedules"), command=self.clear_schedules)
        tools_menu.add_command(label=t("import_sample_data"), command=self.import_sample_data)
        tools_menu.add_command(label=t("backup_database"), command=self.backup_database)
        
        # Language menu
        language_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=t("menu_language"), menu=language_menu)
        language_menu.add_command(label=t("english"), command=lambda: self.change_language("en"))
        language_menu.add_command(label=t("arabic"), command=lambda: self.change_language("ar"))
    
    def change_language(self, language: str):
        """Change the application language"""
        self.localization.set_language(language)
        # Show restart message
        messagebox.showinfo(
            t("menu_language"),
            "Please restart the application for language changes to take effect.\n\n" +
            "يرجى إعادة تشغيل التطبيق لتطبيق تغييرات اللغة."
        )

    def manage_teachers(self):
        """Open teacher management window"""
        self.open_data_management_window(t("manage_teachers"), "teachers", ["name", "availability_json"])

    def manage_classes(self):
        """Open class management window"""
        self.open_data_management_window(t("manage_classes"), "classes", ["name", "grade_level"])

    def manage_subjects(self):
        """Open subject management window"""
        self.open_data_management_window(t("manage_subjects"), "subjects", ["name", "needs_lab"])

    def manage_rooms(self):
        """Open room management window"""
        self.open_data_management_window(t("manage_rooms"), "rooms", ["name", "is_lab"])

    def manage_lessons(self):
        """Open lesson requirements management window"""
        print("Lesson management window - to be implemented")

    def open_data_management_window(self, title, table_name, columns):
        """Open a generic data management window"""
        window = tk.Toplevel(self)
        window.title(title)
        window.geometry("600x400")
        
        # Apply RTL layout if Arabic
        font_family = self.localization.get_font_family()
        if self.localization.is_rtl():
            window.option_add('*TLabel*font', font_family)
        
        # Create treeview for data display
        tree_frame = ttk.Frame(window)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='tree headings')
        tree.pack(fill=tk.BOTH, expand=True)
        
        # Configure columns
        tree.heading('#0', text=t("id"))
        tree.column('#0', width=50)
        for col in columns:
            # Try to get localized column name, fallback to formatted column name
            col_text = t(col) if col in ['name', 'subject', 'teacher', 'class', 'room'] else col.replace('_', ' ').title()
            tree.heading(col, text=col_text)
            tree.column(col, width=150)
        
        # Load data
        self.refresh_data_tree(tree, table_name, columns)
        
        # Buttons frame
        buttons_frame = ttk.Frame(window)
        buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(buttons_frame, text=t("add"), command=lambda: self.add_record(table_name, columns, tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text=t("edit"), command=lambda: self.edit_record(table_name, columns, tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text=t("delete"), command=lambda: self.delete_record(table_name, tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text=t("refresh"), command=lambda: self.refresh_data_tree(tree, table_name, columns)).pack(side=tk.LEFT, padx=5)

    def refresh_data_tree(self, tree, table_name, columns):
        """Refresh the data in a treeview"""
        for item in tree.get_children():
            tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(f"SELECT id, {', '.join(columns)} FROM {table_name}")
        for row in cursor.fetchall():
            tree.insert('', 'end', text=row[0], values=row[1:])
        conn.close()

    def add_record(self, table_name, columns, tree):
        """Add a new record"""
        self.edit_record_window(table_name, columns, tree, None)

    def edit_record(self, table_name, columns, tree):
        """Edit selected record"""
        selection = tree.selection()
        if not selection:
            return
        
        item = tree.item(selection[0])
        record_id = item['text']
        values = item['values']
        
        self.edit_record_window(table_name, columns, tree, record_id, values)

    def edit_record_window(self, table_name, columns, tree, record_id=None, values=None):
        """Open edit/add record window"""
        window = tk.Toplevel(self)
        window.title(f"{t('edit') if record_id else t('add')} {table_name.title()[:-1]}")
        window.geometry("400x300")
        
        entries = {}
        for i, col in enumerate(columns):
            ttk.Label(window, text=col.replace('_', ' ').title() + ":").pack(pady=5)
            entry = ttk.Entry(window, width=30)
            entry.pack(pady=5)
            if values:
                entry.insert(0, str(values[i]))
            entries[col] = entry
        
        def save_record():
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            field_values = [entries[col].get() for col in columns]
            
            if record_id:  # Edit
                placeholders = ', '.join([f"{col} = ?" for col in columns])
                cursor.execute(f"UPDATE {table_name} SET {placeholders} WHERE id = ?", field_values + [record_id])
            else:  # Add
                placeholders = ', '.join(['?' for _ in columns])
                cursor.execute(f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})", field_values)
            
            conn.commit()
            conn.close()
            
            self.refresh_data_tree(tree, table_name, columns)
            self.load_initial_data()  # Refresh main window data
            window.destroy()
        
        ttk.Button(window, text=t("save"), command=save_record).pack(pady=10)

    def delete_record(self, table_name, tree):
        """Delete selected record"""
        selection = tree.selection()
        if not selection:
            return
        
        record_id = tree.item(selection[0])['text']
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM {table_name} WHERE id = ?", (record_id,))
        conn.commit()
        conn.close()
        
        self.refresh_data_tree(tree, table_name, tree.cget('columns'))
        self.load_initial_data()  # Refresh main window data

    def export_pdf(self):
        """Export current view to PDF"""
        try:
            from ..utils.export import export_schedule_to_pdf
            
            selected_item = self.item_selector.get()
            view_type = self.view_var.get()
            
            if not selected_item:
                print(t("please_select_item"))
                return
            
            # Prepare data for PDF export
            days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
            times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
            
            schedule_data = self.load_schedule_data(selected_item, view_type)
            
            # Create table data
            table_data = [[t('time_slot')] + days]
            
            for r, time in enumerate(times):
                row = [time]
                for c, day in enumerate(days):
                    slot_data = schedule_data.get((c, r), None)
                    if slot_data:
                        if view_type == "Classes":
                            cell_text = f"{slot_data['subject']}\n{slot_data['teacher']}\n{slot_data['room']}"
                        else:
                            cell_text = f"{slot_data['subject']}\n{slot_data['class']}\n{slot_data['room']}"
                    else:
                        cell_text = ""
                    row.append(cell_text)
                table_data.append(row)
            
            filename = f"{selected_item.replace(' ', '_')}_{view_type.lower()}_timetable.pdf"
            export_schedule_to_pdf(table_data, filename)
            print(f"{t('pdf_exported')} {filename}")
            
        except Exception as e:
            print(f"Error exporting PDF: {e}")

    def export_excel(self):
        """Export all schedules to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
            times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
            
            # Create sheets for each class
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM classes ORDER BY name")
            classes = [row[0] for row in cursor.fetchall()]
            
            for class_name in classes:
                ws = wb.create_sheet(title=class_name)
                
                # Headers
                ws.cell(row=1, column=1, value=t("time_slot"))
                for c, day in enumerate(days):
                    ws.cell(row=1, column=c + 2, value=day)
                
                # Style headers
                for col in range(1, len(days) + 2):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Get schedule data for this class
                schedule_data = self.load_schedule_data(class_name, "Classes")
                
                # Fill in timetable
                for r, time in enumerate(times):
                    ws.cell(row=r + 2, column=1, value=time)
                    
                    for c, day in enumerate(days):
                        slot_data = schedule_data.get((c, r), None)
                        if slot_data:
                            cell_text = f"{slot_data['subject']}\n{slot_data['teacher']}\n{slot_data['room']}"
                        else:
                            cell_text = ""
                        
                        cell = ws.cell(row=r + 2, column=c + 2, value=cell_text)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        if slot_data:
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                # Adjust column widths
                for col in range(1, len(days) + 2):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
                
                # Adjust row heights
                for row in range(2, len(times) + 2):
                    ws.row_dimensions[row].height = 60
            
            conn.close()
            
            # Save file
            filename = "school_timetable.xlsx"
            wb.save(filename)
            print(f"{t('export_to_excel')} {filename}")
            
        except ImportError:
            print(t("openpyxl_not_installed"))
        except Exception as e:
            print(f"Error exporting Excel: {e}")

    def manage_teacher_preferences(self):
        """Open teacher preferences management window"""
        window = tk.Toplevel(self)
        window.title(t("teacher_preferences"))
        window.geometry("800x600")
        
        # Main frame
        main_frame = ttk.Frame(window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Instructions
        ttk.Label(main_frame, text="Set teacher preferences for classes (1-5 scale, 5=highest preference):", 
                 font=('Helvetica', 12, 'bold')).pack(pady=10)
        
        # Create treeview
        columns = (t('teacher'), t('class'), t('preference'))
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load preferences
        def refresh_preferences():
            for item in tree.get_children():
                tree.delete(item)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT t.name, c.name, tp.preference_score, tp.id
                FROM teacher_preferences tp
                JOIN teachers t ON tp.teacher_id = t.id
                JOIN classes c ON tp.class_id = c.id
                ORDER BY t.name, c.name
            """)
            
            for row in cursor.fetchall():
                tree.insert('', 'end', values=row[:3], tags=(row[3],))
            conn.close()
        
        refresh_preferences()
        
        # Buttons frame
        btn_frame = ttk.Frame(window)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(btn_frame, text=t("add_preference"), 
                  command=lambda: self.add_teacher_preference(tree, refresh_preferences)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("edit_selected"), 
                  command=lambda: self.edit_teacher_preference(tree, refresh_preferences)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("delete_selected"), 
                  command=lambda: self.delete_teacher_preference(tree, refresh_preferences)).pack(side=tk.LEFT, padx=5)

    def add_teacher_preference(self, tree, refresh_callback):
        """Add new teacher preference"""
        window = tk.Toplevel(self)
        window.title(t("add_teacher_preference"))
        window.geometry("400x250")
        
        ttk.Label(window, text=f"{t('teacher')}:").pack(pady=5)
        teacher_var = tk.StringVar()
        teacher_combo = ttk.Combobox(window, textvariable=teacher_var, state="readonly")
        teacher_combo.pack(pady=5)
        
        ttk.Label(window, text=f"{t('class')}:").pack(pady=5)
        class_var = tk.StringVar()
        class_combo = ttk.Combobox(window, textvariable=class_var, state="readonly")
        class_combo.pack(pady=5)
        
        ttk.Label(window, text=f"{t('preference')} (1-5):").pack(pady=5)
        score_var = tk.StringVar(value="3")
        score_spin = ttk.Spinbox(window, from_=1, to=5, textvariable=score_var, width=10)
        score_spin.pack(pady=5)
        
        # Load teachers and classes
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT name FROM teachers ORDER BY name")
        teachers = [row[0] for row in cursor.fetchall()]
        teacher_combo['values'] = teachers
        
        cursor.execute("SELECT name FROM classes ORDER BY name")
        classes = [row[0] for row in cursor.fetchall()]
        class_combo['values'] = classes
        
        conn.close()
        
        def save_preference():
            if not teacher_var.get() or not class_var.get():
                return
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get IDs
            cursor.execute("SELECT id FROM teachers WHERE name = ?", (teacher_var.get(),))
            teacher_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT id FROM classes WHERE name = ?", (class_var.get(),))
            class_id = cursor.fetchone()[0]
            
            # Insert preference
            cursor.execute("""
                INSERT OR REPLACE INTO teacher_preferences (teacher_id, class_id, preference_score)
                VALUES (?, ?, ?)
            """, (teacher_id, class_id, int(score_var.get())))
            
            conn.commit()
            conn.close()
            
            refresh_callback()
            window.destroy()
        
        ttk.Button(window, text=t("save"), command=save_preference).pack(pady=20)

    def edit_teacher_preference(self, tree, refresh_callback):
        """Edit selected teacher preference"""
        selection = tree.selection()
        if not selection:
            return
        
        item = tree.item(selection[0])
        pref_id = item['tags'][0]
        values = item['values']
        
        window = tk.Toplevel(self)
        window.title(t("edit_teacher_preference"))
        window.geometry("400x200")
        
        ttk.Label(window, text=f"{t('teacher')}: {values[0]}").pack(pady=5)
        ttk.Label(window, text=f"{t('class')}: {values[1]}").pack(pady=5)
        
        ttk.Label(window, text=f"{t('preference')} (1-5):").pack(pady=5)
        score_var = tk.StringVar(value=str(values[2]))
        score_spin = ttk.Spinbox(window, from_=1, to=5, textvariable=score_var, width=10)
        score_spin.pack(pady=5)
        
        def save_preference():
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("UPDATE teacher_preferences SET preference_score = ? WHERE id = ?", 
                         (int(score_var.get()), pref_id))
            conn.commit()
            conn.close()
            
            refresh_callback()
            window.destroy()
        
        ttk.Button(window, text=t("save"), command=save_preference).pack(pady=20)

    def delete_teacher_preference(self, tree, refresh_callback):
        """Delete selected teacher preference"""
        selection = tree.selection()
        if not selection:
            return
        
        item = tree.item(selection[0])
        pref_id = item['tags'][0]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM teacher_preferences WHERE id = ?", (pref_id,))
        conn.commit()
        conn.close()
        
        refresh_callback()

    def manage_teacher_availability(self):
        """Open teacher availability management window"""
        window = tk.Toplevel(self)
        window.title(t("teacher_availability"))
        window.geometry("900x700")
        
        # Teacher selection
        top_frame = ttk.Frame(window)
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(top_frame, text=f"{t('select_teacher')}:", font=('Helvetica', 12, 'bold')).pack(side=tk.LEFT, padx=10)
        teacher_var = tk.StringVar()
        teacher_combo = ttk.Combobox(top_frame, textvariable=teacher_var, state="readonly", width=30)
        teacher_combo.pack(side=tk.LEFT, padx=10)
        
        # Load teachers
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM teachers ORDER BY name")
        teachers = [row[0] for row in cursor.fetchall()]
        teacher_combo['values'] = teachers
        conn.close()
        
        # Availability grid frame
        grid_frame = ttk.Frame(window)
        grid_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create availability grid
        days = [t("monday"), t("tuesday"), t("wednesday"), t("thursday"), t("friday")]
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", 
                "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
        
        # Headers
        ttk.Label(grid_frame, text="Time / Day", font=('Helvetica', 10, 'bold')).grid(row=0, column=0, padx=2, pady=2)
        for i, day in enumerate(days):
            ttk.Label(grid_frame, text=day, font=('Helvetica', 10, 'bold')).grid(row=0, column=i+1, padx=2, pady=2)
        
        # Time labels and checkboxes
        availability_vars = {}
        for r, time in enumerate(times):
            ttk.Label(grid_frame, text=time, font=('Helvetica', 9)).grid(row=r+1, column=0, padx=2, pady=2, sticky="e")
            for c, day in enumerate(days):
                var = tk.BooleanVar(value=True)  # Default available
                availability_vars[(c, r)] = var
                cb = ttk.Checkbutton(grid_frame, variable=var, text=t("available"))
                cb.grid(row=r+1, column=c+1, padx=2, pady=2)
        
        def load_teacher_availability():
            """Load availability for selected teacher"""
            if not teacher_var.get():
                return
            
            # Reset all to available
            for var in availability_vars.values():
                var.set(True)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT availability_json FROM teachers WHERE name = ?", (teacher_var.get(),))
            result = cursor.fetchone()
            conn.close()
            
            if result and result[0]:
                availability_data = json.loads(result[0])
                for day_str, periods in availability_data.items():
                    day = int(day_str)
                    for period in periods:
                        if (day, period) in availability_vars:
                            availability_vars[(day, period)].set(False)  # Unavailable
        
        def save_teacher_availability():
            """Save availability for selected teacher"""
            if not teacher_var.get():
                return
            
            # Build availability JSON
            unavailable = {}
            for (day, period), var in availability_vars.items():
                if not var.get():  # If not available
                    if day not in unavailable:
                        unavailable[day] = []
                    unavailable[day].append(period)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("UPDATE teachers SET availability_json = ? WHERE name = ?", 
                         (json.dumps(unavailable), teacher_var.get()))
            conn.commit()
            conn.close()
            
            print(f"Availability updated for {teacher_var.get()}")
        
        teacher_combo.bind("<<ComboboxSelected>>", lambda e: load_teacher_availability())
        
        # Buttons
        btn_frame = ttk.Frame(window)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(btn_frame, text=t("save_availability"), command=save_teacher_availability).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("mark_all_available"), 
                  command=lambda: [var.set(True) for var in availability_vars.values()]).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("mark_all_unavailable"), 
                  command=lambda: [var.set(False) for var in availability_vars.values()]).pack(side=tk.LEFT, padx=5)

    def manage_lessons(self):
        """Enhanced lesson requirements management"""
        window = tk.Toplevel(self)
        window.title(t("lesson_requirements"))
        window.geometry("800x600")
        
        # Create treeview
        columns = (t('class'), t('subject'), t('lessons_per_week'))
        tree = ttk.Treeview(window, columns=columns, show='headings', height=20)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200)
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        def refresh_lessons():
            for item in tree.get_children():
                tree.delete(item)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT c.name, s.name, l.lessons_per_week, l.id
                FROM lessons l
                JOIN classes c ON l.class_id = c.id
                JOIN subjects s ON l.subject_id = s.id
                ORDER BY c.name, s.name
            """)
            
            for row in cursor.fetchall():
                tree.insert('', 'end', values=row[:3], tags=(row[3],))
            conn.close()
        
        refresh_lessons()
        
        # Buttons
        btn_frame = ttk.Frame(window)
        btn_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Button(btn_frame, text=t("add_requirement"), 
                  command=lambda: self.add_lesson_requirement(tree, refresh_lessons)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("edit_selected"), 
                  command=lambda: self.edit_lesson_requirement(tree, refresh_lessons)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text=t("delete_selected"), 
                  command=lambda: self.delete_lesson_requirement(tree, refresh_lessons)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Auto-Generate", 
                  command=lambda: self.auto_generate_lessons(refresh_lessons)).pack(side=tk.LEFT, padx=5)

    def add_lesson_requirement(self, tree, refresh_callback):
        """Add new lesson requirement"""
        window = tk.Toplevel(self)
        window.title(t("add_lesson_requirement"))
        window.geometry("400x250")
        
        ttk.Label(window, text=f"{t('class')}:").pack(pady=5)
        class_var = tk.StringVar()
        class_combo = ttk.Combobox(window, textvariable=class_var, state="readonly")
        class_combo.pack(pady=5)
        
        ttk.Label(window, text=f"{t('subject')}:").pack(pady=5)
        subject_var = tk.StringVar()
        subject_combo = ttk.Combobox(window, textvariable=subject_var, state="readonly")
        subject_combo.pack(pady=5)
        
        ttk.Label(window, text=f"{t('lessons_per_week')}:").pack(pady=5)
        lessons_var = tk.StringVar(value="3")
        lessons_spin = ttk.Spinbox(window, from_=1, to=10, textvariable=lessons_var, width=10)
        lessons_spin.pack(pady=5)
        
        # Load classes and subjects
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT name FROM classes ORDER BY name")
        classes = [row[0] for row in cursor.fetchall()]
        class_combo['values'] = classes
        
        cursor.execute("SELECT name FROM subjects ORDER BY name")
        subjects = [row[0] for row in cursor.fetchall()]
        subject_combo['values'] = subjects
        
        conn.close()
        
        def save_requirement():
            if not class_var.get() or not subject_var.get():
                return
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get IDs
            cursor.execute("SELECT id FROM classes WHERE name = ?", (class_var.get(),))
            class_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT id FROM subjects WHERE name = ?", (subject_var.get(),))
            subject_id = cursor.fetchone()[0]
            
            # Insert requirement
            cursor.execute("""
                INSERT OR REPLACE INTO lessons (class_id, subject_id, lessons_per_week)
                VALUES (?, ?, ?)
            """, (class_id, subject_id, int(lessons_var.get())))
            
            conn.commit()
            conn.close()
            
            refresh_callback()
            window.destroy()
        
        ttk.Button(window, text=t("save"), command=save_requirement).pack(pady=20)

    def edit_lesson_requirement(self, tree, refresh_callback):
        """Edit selected lesson requirement"""
        selection = tree.selection()
        if not selection:
            return
        
        item = tree.item(selection[0])
        lesson_id = item['tags'][0]
        values = item['values']
        
        window = tk.Toplevel(self)
        window.title(t("edit_lesson_requirement"))
        window.geometry("400x200")
        
        ttk.Label(window, text=f"{t('class')}: {values[0]}").pack(pady=5)
        ttk.Label(window, text=f"{t('subject')}: {values[1]}").pack(pady=5)
        
        ttk.Label(window, text=f"{t('lessons_per_week')}:").pack(pady=5)
        lessons_var = tk.StringVar(value=str(values[2]))
        lessons_spin = ttk.Spinbox(window, from_=1, to=10, textvariable=lessons_var, width=10)
        lessons_spin.pack(pady=5)
        
        def save_requirement():
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("UPDATE lessons SET lessons_per_week = ? WHERE id = ?", 
                         (int(lessons_var.get()), lesson_id))
            conn.commit()
            conn.close()
            
            refresh_callback()
            window.destroy()
        
        ttk.Button(window, text=t("save"), command=save_requirement).pack(pady=20)

    def delete_lesson_requirement(self, tree, refresh_callback):
        """Delete selected lesson requirement"""
        selection = tree.selection()
        if not selection:
            return
        
        item = tree.item(selection[0])
        lesson_id = item['tags'][0]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM lessons WHERE id = ?", (lesson_id,))
        conn.commit()
        conn.close()
        
        refresh_callback()

    def auto_generate_lessons(self, refresh_callback):
        """Auto-generate standard lesson requirements"""
        result = tk.messagebox.askyesno("Auto-Generate Lessons", 
                                       "This will generate standard lesson requirements for all classes based on their grade level. Continue?")
        if not result:
            return
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Clear existing lessons
        cursor.execute("DELETE FROM lessons")
        
        # Standard lesson counts by subject name
        lesson_counts = {
            'Mathematics': 5, 'Algebra': 5, 'Geometry': 4, 'Calculus': 4, 'Statistics': 3,
            'English': 4, 'Literature': 4, 'Advanced English': 3, 'Philosophy': 2, 'Creative Writing': 2,
            'Science': 4, 'Biology': 4, 'Physics': 4, 'Chemistry': 4, 'Advanced Physics': 3,
            'Advanced Biology': 3, 'Advanced Chemistry': 3, 'Environmental Science': 3, 'Organic Chemistry': 3,
            'History': 3, 'World History': 3, 'Modern History': 3, 'Economics': 3, 'Political Science': 2,
            'Geography': 3, 'Physical Education': 2, 'Health': 2, 'Psychology': 2, 'Sociology': 2,
            'Art': 2, 'Music': 2, 'Drama': 2, 'Computer Science': 3, 'Engineering': 3
        }
        
        # Subject mapping by grade
        subjects_by_grade = {
            9: ['Mathematics', 'English', 'Science', 'History', 'Geography', 'Physical Education', 'Art'],
            10: ['Algebra', 'Literature', 'Biology', 'World History', 'Chemistry', 'Physical Education', 'Music'],
            11: ['Geometry', 'Advanced English', 'Physics', 'Modern History', 'Environmental Science', 'Health', 'Drama'],
            12: ['Calculus', 'Philosophy', 'Advanced Physics', 'Economics', 'Advanced Chemistry', 'Psychology', 'Computer Science'],
            13: ['Statistics', 'Creative Writing', 'Advanced Biology', 'Political Science', 'Organic Chemistry', 'Sociology', 'Engineering']
        }
        
        # Get all classes and subjects
        cursor.execute("SELECT id, name, grade_level FROM classes")
        classes = cursor.fetchall()
        
        cursor.execute("SELECT id, name FROM subjects")
        subjects = {name: id for id, name in cursor.fetchall()}
        
        # Generate lessons for each class
        for class_id, class_name, grade_level in classes:
            if grade_level in subjects_by_grade:
                for subject_name in subjects_by_grade[grade_level]:
                    if subject_name in subjects:
                        subject_id = subjects[subject_name]
                        lessons_per_week = lesson_counts.get(subject_name, 3)
                        
                        cursor.execute("""
                            INSERT INTO lessons (class_id, subject_id, lessons_per_week)
                            VALUES (?, ?, ?)
                        """, (class_id, subject_id, lessons_per_week))
        
        conn.commit()
        conn.close()
        
        refresh_callback()
        tk.messagebox.showinfo(t("success"), t("lesson_requirements_generated"))

    def manage_scheduling_rules(self):
        """Manage scheduling rules and constraints"""
        window = tk.Toplevel(self)
        window.title("Scheduling Rules")
        window.geometry("600x500")
        
        notebook = ttk.Notebook(window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Hard constraints tab
        hard_frame = ttk.Frame(notebook)
        notebook.add(hard_frame, text="Hard Constraints")
        
        ttk.Label(hard_frame, text="Hard Constraints (Cannot be violated):", 
                 font=('Helvetica', 12, 'bold')).pack(pady=10)
        
        constraints_text = """
        ✓ No teacher can be in two places at once
        ✓ No class can have two lessons simultaneously  
        ✓ No room can host multiple lessons at once
        ✓ Teachers must be available during assigned times
        ✓ Lab subjects must use appropriate lab rooms
        ✓ Exact number of lessons per subject must be scheduled
        """
        
        ttk.Label(hard_frame, text=constraints_text, justify=tk.LEFT).pack(pady=10, padx=20)
        
        # Soft constraints tab
        soft_frame = ttk.Frame(notebook)
        notebook.add(soft_frame, text="Soft Constraints")
        
        ttk.Label(soft_frame, text="Soft Constraints (Optimization goals):", 
                 font=('Helvetica', 12, 'bold')).pack(pady=10)
        
        # Teacher preferences weight
        pref_frame = ttk.Frame(soft_frame)
        pref_frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(pref_frame, text="Teacher Preferences Weight:").pack(side=tk.LEFT)
        pref_var = tk.StringVar(value="1.0")
        ttk.Scale(pref_frame, from_=0.0, to=5.0, variable=pref_var, orient=tk.HORIZONTAL).pack(side=tk.RIGHT, padx=10)
        
        # Gap minimization
        gap_frame = ttk.Frame(soft_frame)
        gap_frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(gap_frame, text="Minimize Teacher Gaps:").pack(side=tk.LEFT)
        gap_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(gap_frame, variable=gap_var).pack(side=tk.RIGHT)
        
        # Consecutive lessons limit
        consec_frame = ttk.Frame(soft_frame)
        consec_frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(consec_frame, text="Max Consecutive Same Subject:").pack(side=tk.LEFT)
        consec_var = tk.StringVar(value="2")
        ttk.Spinbox(consec_frame, from_=1, to=5, textvariable=consec_var, width=5).pack(side=tk.RIGHT)

    def manage_constraints(self):
        """Manage specific constraint settings"""
        window = tk.Toplevel(self)
        window.title("Constraint Settings")
        window.geometry("500x400")
        
        ttk.Label(window, text="Constraint Configuration", 
                 font=('Helvetica', 14, 'bold')).pack(pady=20)
        
        # Solver timeout
        timeout_frame = ttk.Frame(window)
        timeout_frame.pack(fill=tk.X, padx=20, pady=10)
        ttk.Label(timeout_frame, text="Solver Timeout (seconds):").pack(side=tk.LEFT)
        timeout_var = tk.StringVar(value="30")
        ttk.Entry(timeout_frame, textvariable=timeout_var, width=10).pack(side=tk.RIGHT)
        
        # Max periods per day
        periods_frame = ttk.Frame(window)
        periods_frame.pack(fill=tk.X, padx=20, pady=10)
        ttk.Label(periods_frame, text="Periods per Day:").pack(side=tk.LEFT)
        periods_var = tk.StringVar(value="8")
        ttk.Spinbox(periods_frame, from_=6, to=10, textvariable=periods_var, width=5).pack(side=tk.RIGHT)
        
        # Working days
        days_frame = ttk.Frame(window)
        days_frame.pack(fill=tk.X, padx=20, pady=10)
        ttk.Label(days_frame, text="Working Days per Week:").pack(side=tk.LEFT)
        days_var = tk.StringVar(value="5")
        ttk.Spinbox(days_frame, from_=4, to=6, textvariable=days_var, width=5).pack(side=tk.RIGHT)

    def manage_time_settings(self):
        """Manage time slot settings"""
        window = tk.Toplevel(self)
        window.title("Time Settings")
        window.geometry("600x500")
        
        ttk.Label(window, text="Time Slot Configuration", 
                 font=('Helvetica', 14, 'bold')).pack(pady=20)
        
        # Current time slots display
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", "11:00-12:00", 
                "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00"]
        
        ttk.Label(window, text="Current Time Slots:", font=('Helvetica', 12, 'bold')).pack(pady=10)
        
        for time in times:
            ttk.Label(window, text=f"Period {times.index(time)+1}: {time}").pack()
        
        ttk.Label(window, text="\nTime slots are currently fixed but can be modified in the code.", 
                 font=('Helvetica', 10, 'italic')).pack(pady=20)

    def show_database_stats(self):
        """Show database statistics"""
        window = tk.Toplevel(self)
        window.title(t("database_statistics"))
        window.geometry("500x400")
        
        # Get statistics
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        stats = {}
        tables = ['teachers', 'classes', 'subjects', 'rooms', 'lessons', 'teacher_preferences', 'schedules']
        
        for table in tables:
            cursor.execute(f"SELECT COUNT(*) FROM {table}")
            stats[table] = cursor.fetchone()[0]
        
        # Get additional stats
        cursor.execute("SELECT COUNT(*) FROM subjects WHERE needs_lab = 1")
        lab_subjects = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM rooms WHERE is_lab = 1")  
        lab_rooms = cursor.fetchone()[0]
        
        cursor.execute("SELECT SUM(lessons_per_week) FROM lessons")
        total_lessons_needed = cursor.fetchone()[0] or 0
        
        conn.close()
        
        # Display stats
        ttk.Label(window, text=t("database_statistics"), font=('Helvetica', 14, 'bold')).pack(pady=20)
        
        stats_text = f"""
        {t('teachers')}: {stats['teachers']}
        {t('classes')}: {stats['classes']}
        {t('subjects')}: {stats['subjects']} ({lab_subjects} {t('require_labs')})
        {t('rooms')}: {stats['rooms']} ({lab_rooms} {t('are_labs')})
        
        {t('lesson_requirements')}: {stats['lessons']}
        {t('total_lessons_needed')}: {total_lessons_needed}
        {t('scheduled_lessons')}: {stats['schedules']}
        
        {t('teacher_preferences')}: {stats['teacher_preferences']}
        """
        
        ttk.Label(window, text=stats_text, justify=tk.LEFT, font=('Helvetica', 11)).pack(pady=20)
        
        # Utilization info
        if total_lessons_needed > 0:
            utilization = (stats['schedules'] / total_lessons_needed) * 100
            ttk.Label(window, text=f"{t('utilization_percentage')}: {utilization:.1f}%", 
                     font=('Helvetica', 12, 'bold')).pack(pady=10)

    def clear_schedules(self):
        """Clear all generated schedules"""
        result = tk.messagebox.askyesno("Clear Schedules", 
                                       "This will delete all generated schedules. Continue?")
        if result:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM schedules WHERE is_locked = 0")
            conn.commit()
            conn.close()
            
            self.draw_timetable()
            tk.messagebox.showinfo(t("success"), t("all_schedules_cleared"))

    def import_sample_data(self):
        """Import fresh sample data"""
        result = tk.messagebox.askyesno("Import Sample Data", 
                                       "This will replace all current data with fresh sample data. Continue?")
        if result:
            from ..database.database_setup import add_sample_data, create_connection
            conn = create_connection()
            if conn:
                add_sample_data(conn)
                conn.close()
                self.load_initial_data()
                self.draw_timetable()
                tk.messagebox.showinfo(t("success"), t("sample_data_imported"))

    def backup_database(self):
        """Create a backup of the database"""
        import shutil
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"school_timetable_backup_{timestamp}.db"
        
        try:
            shutil.copy2("school_timetable.db", backup_name)
            tk.messagebox.showinfo(t("backup_created"), f"{t('database_backed_up')}: {backup_name}")
        except Exception as e:
            tk.messagebox.showerror(t("backup_failed"), f"{t('failed_to_create_backup')}: {e}")

if __name__ == "__main__":
    app = TimetableApp()
    app.mainloop()
